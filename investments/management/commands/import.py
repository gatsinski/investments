import re
from datetime import datetime
from decimal import Decimal

import openpyxl
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.utils.timezone import make_aware

from investments.contrib.brokers.models import Broker
from investments.contrib.payments.models import DividendPayment
from investments.contrib.positions.models import Position
from investments.contrib.securities.models import Stock

UserModel = get_user_model()


class Command(BaseCommand):
    help = "Imports investment data from xls"

    def add_arguments(self, parser):
        parser.add_argument("file", nargs=1, type=str)
        parser.add_argument("user", nargs="?", type=str)
        parser.add_argument("broker", nargs="?", type=str)

    def handle(self, *args, **options):
        filename = options.get("file")[0]

        user_email = options.get("user")
        user = self.get_user(user_email)

        if not user:
            return

        broker_name = options.get("broker")
        broker = self.get_broker(broker_name, user)

        if not broker:
            return

        workbook = openpyxl.load_workbook(
            filename=filename, read_only=True, data_only=True
        )

        self.handle_positions(workbook, user, broker)
        self.handle_dividends(workbook)

        workbook.close()

        self.write_success("Successfully imported all data")

    def get_user(self, user_email):
        if user_email:
            user = UserModel.objects.filter(email=user_email).first()

            if user:
                return user

            self.write_error(f"Unable to find the user with email {user_email}.")
        else:
            self.write_warning("No user provided. Using the first available user.")

            user = UserModel.objects.last()

            if user:
                return user

            self.write_error("Unable to find user.")

    def get_broker(self, broker_name, user):
        if broker_name:
            broker = Broker.objects.filter(user=user, name=broker_name).first()

            if broker:
                return broker

            self.write_error(f"Unable to find broker with name {broker_name}.")
        else:
            self.write_warning(
                f"No broker provided. Using the first broker belonging to {user.email}."
            )

            broker = Broker.objects.filter(user=user).first()

            if broker:
                return broker

            self.write_error(f"Unable to find any broker belonging to {user.email}.")

    def handle_dividends(self, workbook):
        worksheet = workbook["Dividends"]

        for row in worksheet.iter_rows(min_row=2):
            date = row[0]
            aware_date = make_aware(datetime.strptime(date.value, "%d/%m/%Y"))

            received_dividend = row[2]
            withheld_tax_rate = Decimal(re.findall(r"\d+", row[3].value)[0])
            withheld_tax = Decimal(str(row[4].value))

            position_id = row[5]

            position = Position.objects.filter(position_id=position_id.value).first()

            if not position:
                self.write_error(f"Failed to find {position_id.value}. Skipping.")
                continue

            payment, created = DividendPayment.objects.get_or_create(
                position=position,
                amount=received_dividend.value,
                recorded_on=aware_date,
                defaults={
                    "withheld_tax": withheld_tax,
                    "withheld_tax_rate": withheld_tax_rate,
                },
            )

            if created:
                self.write_success(
                    f"Successfully created payment for {position.position_id} ({position.security.name}) "
                    f"with received amount {payment.amount} recorded on {payment.recorded_on} with id {payment.uuid}."
                )
            else:
                self.write_warning(
                    f"Found existing payment for {position.position_id} ({position.security.name}) with id {payment.uuid}"
                )

                if not payment.withheld_tax or not payment.withheld_tax_rate:
                    payment.withheld_tax = withheld_tax
                    payment.withheld_tax_rate = withheld_tax_rate
                    payment.save(update_fields=("withheld_tax", "withheld_tax_rate"))

                    self.write_success(
                        f"The existing payment for {position.position_id} ({position.security.name}) with id {payment.uuid} "
                        f"had no tax data. It was successfully updated with witheld tax {withheld_tax} and rate {withheld_tax_rate}%."
                    )

    def handle_positions(self, workbook, user, broker):
        worksheet = workbook["Account Activity"]

        for row in worksheet.iter_rows(min_row=2):
            activity = {
                "date": make_aware(
                    datetime.strptime(row[0].value, "%d/%m/%Y %H:%M:%S")
                ),
                "type": row[1].value,
                "symbol": row[2].value.split("/")[0] if row[2].value else "",
                "amount": Decimal(str(row[3].value)),
                "units": row[4].value,
                "position_id": row[8].value,
                "asset_type": row[9].value,
            }

            if not activity["position_id"]:
                continue

            position = Position.objects.filter(
                position_id=activity["position_id"]
            ).first()

            if activity["type"] == "Open Position":
                self.handle_opened_positions(activity, position, user, broker)
            elif activity["type"] == "Position closed":
                self.handle_closed_positions(activity, position, user, broker)

    def handle_opened_positions(self, activity, position, user, broker):
        if position:
            self.write_warning(
                f"Position {position.position_id} already exists. Skipping"
            )
            return

        if activity["asset_type"] != "Stocks":
            self.write_warning(
                f"Encountered an asset of type {activity['asset_type']}. Skipping"
            )
            return

        stock = Stock.objects.filter(symbol=activity["symbol"], user=user).first()

        if not stock:
            stock = Stock.objects.filter(
                aliases__contains=activity["symbol"], user=user
            ).first()

        if not stock:
            self.write_error(
                f"Unable to find stock with symbol {activity['symbol']} belonging to user {user.email}. Skipping"
            )
            return

        position = Position.objects.create(
            position_id=activity["position_id"],
            units=activity["units"],
            open_price=activity["amount"] / Decimal(activity["units"]),
            security=stock,
            broker=broker,
            opened_at=activity["date"],
        )

        self.write_success(
            f"Successfully created position with id {position.position_id} - Buy {position.units} {position.security.name} at {position.open_price}"
        )

    def handle_closed_positions(self, activity, position, user, broker):
        if not position:
            self.write_error(f"Failed to find {activity['position_id']}. Skipping.")
            return

        if position.is_closed:
            self.write_error(
                f"Position {activity['position_id']} is already closed. Skipping."
            )
            return

        units = Decimal(activity["units"])

        position.close_price = activity["amount"] / units
        position.closed_at = activity["date"]

        position.save(update_fields=["close_price", "closed_at"])

        self.write_success(
            f"Successfully closed position with id {position.position_id} - {position.units} {position.security.name} at {position.close_price}"
        )

    def write_success(self, message):
        self.stdout.write(self.style.SUCCESS(message))

    def write_warning(self, message):
        self.stdout.write(self.style.WARNING(message))

    def write_error(self, message):
        self.stdout.write(self.style.ERROR(message))
